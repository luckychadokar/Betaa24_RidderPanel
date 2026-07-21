from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from .models import Rider, RiderTraining, RiderDocument, BankDetail, WalletTransaction
from .forms import RiderForm, TrainingForm, DocumentForm, BankDetailForm, ExcelUploadForm
import openpyxl
from io import BytesIO


@login_required
def rider_list(request):
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    riders = Rider.objects.all()
    if q:
        riders = riders.filter(Q(name__icontains=q) | Q(mobile__icontains=q) | Q(city__icontains=q))
    if status:
        riders = riders.filter(status=status)
    dummy = Rider()
    return render(request, 'riders/list.html', {
        'riders': riders, 'q': q, 'status': status, 'rider': dummy
    })


@login_required
def rider_add(request):
    form = RiderForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        rider = form.save()
        RiderTraining.objects.get_or_create(rider=rider)
        RiderDocument.objects.get_or_create(rider=rider)
        messages.success(request, f'Rider {rider.name} added successfully!')
        return redirect('rider_list')
    return render(request, 'riders/form.html', {'form': form, 'title': 'Add Rider'})


@login_required
def rider_edit(request, pk):
    rider = get_object_or_404(Rider, pk=pk)
    form = RiderForm(request.POST or None, request.FILES or None, instance=rider)
    if form.is_valid():
        form.save()
        messages.success(request, 'Rider updated!')
        return redirect('rider_detail', pk=pk)
    return render(request, 'riders/form.html', {'form': form, 'title': 'Edit Rider', 'rider': rider})


@login_required
def rider_detail(request, pk):
    rider = get_object_or_404(Rider, pk=pk)
    training = getattr(rider, 'training', None)
    docs = getattr(rider, 'documents', None)
    tasks = rider.tasks.all()[:10]
    txns = rider.wallet_transactions.all()[:10]
    doc_status = []
    if docs:
        doc_status = [
            ('Aadhaar', docs.aadhaar_status),
            ('PAN', docs.pan_status),
            ('Driving License', docs.dl_status),
            ('Overall', docs.overall_status),
        ]
    wallet_rows = [
        ('Total Earned', rider.total_earnings, 'text-g'),
        ('Total Paid', rider.total_paid, 'text-g'),
        ('Pending', rider.pending_amount, 'text-r'),
    ]
    return render(request, 'riders/detail.html', {
        'rider': rider, 'training': training, 'docs': docs,
        'tasks': tasks, 'txns': txns, 'doc_status': doc_status,
        'wallet_rows': wallet_rows,
    })


@login_required
def rider_delete(request, pk):
    rider = get_object_or_404(Rider, pk=pk)
    if request.method == 'POST':
        rider.delete()
        messages.success(request, 'Rider deleted.')
        return redirect('rider_list')
    return render(request, 'riders/confirm_delete.html', {'obj': rider, 'type': 'Rider'})


@login_required
def training_update(request, pk):
    rider = get_object_or_404(Rider, pk=pk)
    training, _ = RiderTraining.objects.get_or_create(rider=rider)
    form = TrainingForm(request.POST or None, instance=training)
    if form.is_valid():
        form.save()
        messages.success(request, 'Training status updated!')
        return redirect('rider_detail', pk=pk)
    return render(request, 'riders/training_form.html', {'form': form, 'rider': rider})


@login_required
def document_update(request, pk):
    rider = get_object_or_404(Rider, pk=pk)
    doc, _ = RiderDocument.objects.get_or_create(rider=rider)
    form = DocumentForm(request.POST or None, request.FILES or None, instance=doc)
    if form.is_valid():
        form.save()
        messages.success(request, 'Documents updated!')
        return redirect('rider_detail', pk=pk)
    return render(request, 'riders/doc_form.html', {'form': form, 'rider': rider})


@login_required
def excel_import(request):
    form = ExcelUploadForm(request.POST or None, request.FILES or None)
    imported, errors, skipped = [], [], []

    COLUMN_MAP = {
        'beta_code': ['account id', 'accountid', 'id', 'beta code', 'beta_code', 'app id', 'rider id'],
        'name': ['name', 'rider name', 'full name', 'rider_name', 'fullname', 'riders name'],
        'mobile': ['mobile', 'phone', 'mobile number', 'mob number', 'mob. number',
                   'mobile no', 'phone number', 'contact', 'contact number',
                   'mobile_number', 'phone_number', 'number', 'number ', 'mob no'],
        'gender': ['gender', 'sex'],
        'email': ['email', 'email id', 'email address', 'mail', 'e-mail'],
        'state': ['state'],
        'city': ['city', 'location', 'town', 'area'],
        'vehicle_type': ['vehicle_type', 'vehicle', 'vehicle type', 'bike type', 'vehicletype'],
        'status': ['status', 'rider status', 'state status'],
        'joining_date': ['joining_date', 'joining date', 'join date', 'date of joining',
                         'doj', 'joined on', 'date & time', 'activation date', 'date'],
        'address': ['address', 'full address', 'residential address'],
        'emergency_contact': ['emergency_contact', 'emergency contact', 'emergency number',
                              'alternate number', 'alternate contact'],
    }

    def normalize(s):
        if s is None:
            return ''
        s = str(s).strip().lower()
        s = s.replace('.', '').replace('_', ' ').replace('-', ' ')
        s = ' '.join(s.split())
        return s

    def build_header_index(headers):
        norm_headers = [normalize(h) for h in headers]
        field_index = {}
        for field, variations in COLUMN_MAP.items():
            norm_variations = [normalize(v) for v in variations]
            for idx, h in enumerate(norm_headers):
                if h in norm_variations:
                    field_index[field] = idx
                    break
        return field_index

    def get_val(row, field_index, field, default=''):
        idx = field_index.get(field)
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        if val is None:
            return default
        return str(val).strip()

    def clean_mobile(raw):
        if not raw:
            return ''
        s = str(raw).strip()
        if s.endswith('.0'):
            s = s[:-2]
        digits = ''.join(ch for ch in s if ch.isdigit())
        if len(digits) > 10:
            digits = digits[-10:]
        return digits

    def map_status(raw):
        if not raw:
            return 'training_pending'
        s = normalize(raw)
        mapping = {
            'active': 'active',
            'new': 'training_pending',
            'under process': 'training_pending',
            'training pending': 'training_pending',
            'pending training': 'training_pending',
            'documents pending': 'documents_pending',
            'docs pending': 'documents_pending',
            'ready': 'ready',
            'ready for activation': 'ready',
            'inactive': 'inactive',
        }
        return mapping.get(s, 'training_pending')

    def find_header_row(ws, max_scan=5):
        all_rows = list(ws.iter_rows(values_only=True))
        for i in range(min(max_scan, len(all_rows))):
            row = all_rows[i]
            headers = [str(h) if h is not None else '' for h in row]
            field_index = build_header_index(headers)
            if 'mobile' in field_index:
                return i, headers, field_index, all_rows
        return None, None, None, all_rows

    if form.is_valid():
        try:
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            header_idx, headers, field_index, all_rows = find_header_row(ws)

            if header_idx is None:
                errors.append('Could not find Mobile/Phone column in first 5 rows.')
            else:
                data_rows = all_rows[header_idx + 1:]
                for row_num, row in enumerate(data_rows, start=header_idx + 2):
                    if row is None or all(c is None for c in row):
                        continue

                    mobile_raw = get_val(row, field_index, 'mobile')
                    mobile = clean_mobile(mobile_raw)

                    if not mobile or len(mobile) != 10:
                        skipped.append({
                            'row': row_num,
                            'reason': f'Invalid/missing mobile ("{mobile_raw}")'
                        })
                        continue

                    name = get_val(row, field_index, 'name') or f'Rider {mobile}'
                    beta_code = get_val(row, field_index, 'beta_code')
                    gender = get_val(row, field_index, 'gender')
                    email = get_val(row, field_index, 'email')
                    state = get_val(row, field_index, 'state') or 'Madhya Pradesh'
                    city_raw = get_val(row, field_index, 'city')
                    city = city_raw if city_raw and city_raw != '-' else 'Bhopal'
                    vehicle = get_val(row, field_index, 'vehicle_type') or 'Bike'
                    status_raw = get_val(row, field_index, 'status')
                    status = map_status(status_raw)
                    address = get_val(row, field_index, 'address')
                    emergency = clean_mobile(get_val(row, field_index, 'emergency_contact'))

                    try:
                        defaults = {
                            'name': name,
                            'email': email,
                            'city': city,
                            'vehicle_type': vehicle,
                            'status': status,
                            'address': address,
                            'emergency_contact': emergency,
                        }
                        if beta_code:
                            defaults['beta_code'] = beta_code
                        if gender:
                            defaults['gender'] = gender
                        if state:
                            defaults['state'] = state

                        rider, created = Rider.objects.update_or_create(
                            mobile=mobile,
                            defaults=defaults
                        )
                        if created:
                            RiderTraining.objects.get_or_create(rider=rider)
                            RiderDocument.objects.get_or_create(rider=rider)
                        imported.append({
                            'name': rider.name,
                            'mobile': rider.mobile,
                            'action': 'Created' if created else 'Updated'
                        })
                    except Exception as e:
                        skipped.append({'row': row_num, 'reason': str(e)})

        except Exception as e:
            errors.append(f'File could not be processed: {e}')

    return render(request, 'riders/excel_import.html', {
        'form': form,
        'imported': imported,
        'errors': errors,
        'skipped': skipped,
    })


@login_required
def excel_export(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Riders'
    ws.append(['ID', 'Name', 'Mobile', 'Email', 'City', 'Vehicle',
               'Status', 'Rating', 'Total Tasks', 'Total Earnings', 'Joined'])
    for r in Rider.objects.all():
        ws.append([r.id, r.name, r.mobile, r.email, r.city, r.vehicle_type,
                   r.status, str(r.rating), r.total_tasks,
                   str(r.total_earnings), str(r.joining_date)])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename=beta24_riders.xlsx'
    return resp


@login_required
def bank_detail_update(request, pk):
    rider = get_object_or_404(Rider, pk=pk)
    bank, _ = BankDetail.objects.get_or_create(rider=rider)
    form = BankDetailForm(request.POST or None, request.FILES or None, instance=bank)
    if form.is_valid():
        form.save()
        messages.success(request, 'Bank details updated!')
        return redirect('rider_detail', pk=pk)
    return render(request, 'riders/bank_form.html', {'form': form, 'rider': rider})


@login_required
def ifsc_lookup(request):
    import requests as req
    ifsc = request.GET.get('ifsc', '').strip().upper()
    if len(ifsc) != 11:
        return JsonResponse({'error': 'Invalid IFSC code length'}, status=400)
    try:
        resp = req.get(f'https://ifsc.razorpay.com/{ifsc}', timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            return JsonResponse({
                'bank_name': data.get('BANK', ''),
                'branch_name': data.get('BRANCH', ''),
                'city': data.get('CITY', ''),
            })
        else:
            return JsonResponse({'error': 'IFSC not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def training_records(request):
    q = request.GET.get('q', '')
    records = RiderTraining.objects.select_related('rider').all().order_by('-updated_at')
    if q:
        records = records.filter(
            Q(rider__name__icontains=q) |
            Q(rider__mobile__icontains=q) |
            Q(location__icontains=q) |
            Q(city__icontains=q)
        )
    return render(request, 'riders/training_records.html', {'records': records, 'q': q})


@login_required
def training_import(request):
    imported, skipped = [], []

    if request.method == 'POST':
        raw_text = request.POST.get('raw_data', '').strip()
        if raw_text:
            lines = [l for l in raw_text.split('\n') if l.strip()]
            for line_num, line in enumerate(lines, start=1):
                parts = [p.strip() for p in line.split('\t')]

                if len(parts) < 3:
                    skipped.append({'row': line_num, 'reason': 'Too few columns', 'raw': line[:60]})
                    continue

                try:
                    date_str = parts[0] if len(parts) > 0 else ''
                    name = parts[1] if len(parts) > 1 else ''
                    mobile_raw = parts[2] if len(parts) > 2 else ''
                    alt_number = parts[3] if len(parts) > 3 and parts[3].lower() != 'na' else ''
                    location = parts[4] if len(parts) > 4 else ''
                    state = parts[5] if len(parts) > 5 else ''
                    city = parts[6] if len(parts) > 6 else ''
                    done_by = parts[7] if len(parts) > 7 else ''
                    remark = parts[8] if len(parts) > 8 else ''

                    mobile = ''.join(ch for ch in mobile_raw if ch.isdigit())[-10:]

                    if not mobile or len(mobile) != 10:
                        skipped.append({
                            'row': line_num,
                            'reason': f'Invalid mobile: {mobile_raw}',
                            'raw': line[:60]
                        })
                        continue

                    rider, created = Rider.objects.get_or_create(
                        mobile=mobile,
                        defaults={
                            'name': name or f'Rider {mobile}',
                            'city': city or 'Bhopal',
                            'status': 'training_pending',
                        }
                    )
                    if created:
                        RiderDocument.objects.get_or_create(rider=rider)

                    remark_lower = remark.strip().lower()
                    if 'done' in remark_lower:
                        status = 'completed'
                    elif remark_lower in ('npc', 'not intrested', 'not interested'):
                        status = 'rejected'
                    elif remark_lower:
                        status = 'pending'
                    else:
                        status = 'pending'

                    from datetime import datetime
                    training_date = None
                    for fmt in ('%d-%b-%Y', '%d-%b-%y', '%d %b %Y', '%d/%m/%Y'):
                        try:
                            training_date = datetime.strptime(date_str, fmt).date()
                            break
                        except (ValueError, TypeError):
                            continue

                    RiderTraining.objects.update_or_create(
                        rider=rider,
                        defaults={
                            'status': status,
                            'alternate_number': alt_number,
                            'location': location,
                            'state': state,
                            'city': city,
                            'training_done_by': done_by,
                            'final_remark': remark,
                            'training_date': training_date,
                        }
                    )

                    imported.append({
                        'name': rider.name,
                        'mobile': rider.mobile,
                        'status': status,
                        'remark': remark,
                    })
                except Exception as e:
                    skipped.append({'row': line_num, 'reason': str(e), 'raw': line[:60]})

    return render(request, 'riders/training_import.html', {
        'imported': imported,
        'skipped': skipped,
    })


@login_required
def beta24_app_import(request):
    imported, skipped = [], []

    if request.method == 'POST':
        raw_text = request.POST.get('raw_data', '').strip()
        if raw_text:
            lines = [l for l in raw_text.split('\n') if l.strip()]
            for line_num, line in enumerate(lines, start=1):
                parts = line.split('\t')
                parts = [p.strip() for p in parts]

                if len(parts) < 7:
                    skipped.append({
                        'row': line_num,
                        'reason': 'Could not parse columns',
                        'raw': line[:60]
                    })
                    continue

                try:
                    app_id = parts[0]
                    beta_code = parts[1]
                    name = parts[2] if parts[2] != '-' else f'Rider {beta_code}'
                    gender = parts[3] if parts[3] != '-' else ''
                    mobile_raw = parts[4]
                    state = parts[5] if len(parts) > 5 else 'Madhya Pradesh'
                    city = parts[6] if len(parts) > 6 else ''
                    app_status = parts[7] if len(parts) > 7 else 'Under Process'
                    reg_date = parts[8] if len(parts) > 8 else ''

                    mobile = ''.join(ch for ch in mobile_raw if ch.isdigit())[-10:]

                    if not mobile or len(mobile) != 10:
                        skipped.append({
                            'row': line_num,
                            'reason': f'Invalid mobile: {mobile_raw}',
                            'raw': line[:60]
                        })
                        continue

                    status_map = {
                        'under process': 'training_pending',
                        'approved': 'active',
                        'active': 'active',
                        'rejected': 'inactive',
                    }
                    status = status_map.get(app_status.strip().lower(), 'training_pending')

                    rider, created = Rider.objects.update_or_create(
                        mobile=mobile,
                        defaults={
                            'name': name,
                            'gender': gender,
                            'city': city if city and city != '-' else 'Bhopal',
                            'state': state if state and state != '-' else 'Madhya Pradesh',
                            'beta_code': beta_code,
                            'app_status': app_status,
                            'registered_at': reg_date,
                            'status': status,
                        }
                    )
                    if created:
                        RiderTraining.objects.get_or_create(rider=rider)
                        RiderDocument.objects.get_or_create(rider=rider)

                    imported.append({
                        'name': rider.name,
                        'mobile': rider.mobile,
                        'beta_code': rider.beta_code,
                        'action': 'Created' if created else 'Updated'
                    })
                except Exception as e:
                    skipped.append({'row': line_num, 'reason': str(e), 'raw': line[:60]})

    return render(request, 'riders/beta24_import.html', {
        'imported': imported,
        'skipped': skipped,
    })